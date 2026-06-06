"""Generate the Excel tutorial notebook.

Creates examples/excel/excel_workflow.ipynb — a workflow demonstrating
condition and abort_condition for conditional step execution.

Usage:
    python examples/excel/_nb_generate.py
"""

from pathlib import Path

import nbformat as nbf

nb = nbf.v4.new_notebook()
nb.cells = []


def code(s):
    nb.cells.append(nbf.v4.new_code_cell(s))


def md(s):
    nb.cells.append(nbf.v4.new_markdown_cell(s))


md("""\
# Excel Workflow Tutorial — Conditional Steps

This notebook demonstrates an ffai workflow with **conditional execution** using Excel:

- `condition` — skip a step unless the expression is true
- `abort_condition` — stop the entire workflow if the expression is true

The workflow is a content review pipeline:

1. **generate** — Write a short paragraph on a topic
2. **score** — Score the paragraph 1-10 (returns JSON)
3. **revise** — Revise the paragraph *(only runs if score < 8)*
4. **safety** — Check if the content is safe *(aborts workflow if flagged)*
5. **summary** — Summarize what happened

## Prerequisites

```bash
pip install ffai-workflow-adapters[excel]
```
""")


code("""\
import os
import sys
from pathlib import Path

_cwd = Path().resolve()
_project_root = _cwd
for _p in [_cwd, *list(_cwd.parents)]:
    if (_p / "pyproject.toml").is_file():
        _project_root = _p
        break
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from dotenv import load_dotenv  # noqa: E402
load_dotenv()  # noqa: E402

import openpyxl  # noqa: E402
from ffai_workflow_adapters import load_workflow_excel, write_workflow_results_excel  # noqa: E402

print(f"Project root: {_project_root}")
""")


md("""\
---
<div class="page-break"></div>

## Section 1: Create the Workflow Workbook

Build an Excel file with 5 steps demonstrating `condition` and `abort_condition`.
""")


code("""\
WORKFLOW_FILE = str(Path(_project_root) / "examples" / "excel" / "tutorial_workflow.xlsx")
RESULTS_FILE = str(Path(_project_root) / "examples" / "excel" / "tutorial_results.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Workflow"

headers = ["name", "prompt", "condition", "abort_condition", "temperature"]
ws.append(headers)

steps = [
    {
        "name": "generate",
        "prompt": "Write a short paragraph (3-4 sentences) about the future of renewable energy.",
        "condition": "",
        "abort_condition": "",
        "temperature": 0.7,
    },
    {
        "name": "score",
        "prompt": (
            "Rate this paragraph on clarity and insight on a scale of 1-10. "
            'Respond with ONLY a JSON object like: {"score": 7, "pass": true} '
            "where pass is true if score >= 8.\\n\\nParagraph: {{generate.response}}"
        ),
        "condition": "",
        "abort_condition": "",
        "temperature": 0.3,
    },
    {
        "name": "revise",
        "prompt": (
            "The paragraph below scored poorly. Rewrite it to be clearer and more insightful.\\n\\n"
            "Original: {{generate.response}}\\n\\n"
            "Score: {{score.response}}"
        ),
        "condition": 'json_get({{score.response}}, "pass") == False',
        "abort_condition": "",
        "temperature": 0.7,
    },
    {
        "name": "safety",
        "prompt": (
            "Review this text for any safety concerns (harmful content, misinformation, etc.). "
            'Respond with ONLY a JSON object: {"safe": true, "flagged": false, "reason": ""}\\n\\n'
            "Text: {{revise.response}}{{generate.response}}"
        ),
        "condition": "",
        "abort_condition": 'json_get({{safety.response}}, "flagged") == True',
        "temperature": 0.1,
    },
    {
        "name": "summary",
        "prompt": (
            "Summarize the content review process that just happened. "
            "Was the content revised? Is it safe? Provide a brief status.\\n\\n"
            "Score: {{score.response}}\\n"
            "Safety: {{safety.response}}"
        ),
        "condition": "",
        "abort_condition": "",
        "temperature": 0.5,
    },
]

for step in steps:
    ws.append([step.get(h, "") for h in headers])

wb.save(WORKFLOW_FILE)
print(f"Created {WORKFLOW_FILE}")
print(f"Steps: {len(steps)}")
print()
for step in steps:
    cond = f" [condition: {step['condition']}]" if step["condition"] else ""
    abort = f" [abort: {step['abort_condition']}]" if step["abort_condition"] else ""
    print(f"  {step['name']}{cond}{abort}")
""")


md("""\
---
<div class="page-break"></div>

## Section 2: Load the Workflow

Load the Excel file into an ffai `WorkflowSpec`.
""")


code("""\
spec = load_workflow_excel(WORKFLOW_FILE, sheet="Workflow", name="excel_conditional")

print(f"Workflow: {spec.name}")
print(f"Steps: {len(spec.prompts)}")
print()
for p in spec.prompts:
    cond = f" [condition: {p.condition}]" if p.condition else ""
    abort = f" [abort: {p.abort_condition}]" if p.abort_condition else ""
    print(f"  {p.name} (temp={p.temperature}){cond}{abort}")
    print(f"    {p.prompt[:90]}{'...' if len(p.prompt) > 90 else ''}")
""")


md("""\
---
<div class="page-break"></div>

## Section 3: Execute the Workflow

Run the workflow through ffai. The `revise` step will be **skipped** if the
score is 8+, and the workflow will **abort** if safety flags the content.
""")


code("""\
import asyncio
import concurrent.futures
from ffai import FFAI
from ffai.Clients.AsyncFFLiteLLMClient import AsyncFFLiteLLMClient

def run_sync(coro):
    try:
        asyncio.get_running_loop()
    except RuntimeError:
        return asyncio.run(coro)
    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
        return pool.submit(asyncio.run, coro).result()

model_string = os.environ.get("FFAI_MODEL_STRING", "mistral/mistral-small-latest")
api_key_env = "MISTRAL_API_KEY" if "mistral" in model_string else "OPENAI_API_KEY"
api_key = os.environ.get(api_key_env, "")

if not api_key:
    print(f"WARNING: {api_key_env} not set. LLM calls will fail.")

client = AsyncFFLiteLLMClient(model_string=model_string, api_key=api_key)
ffai = FFAI(client)

print(f"LLM client: {model_string}")
print("Executing workflow...")
""")


code("""\
result = run_sync(ffai.workflow.execute_workflow(spec))

print(f"Completed: {result.success_count} succeeded, {result.failed_count} failed, {result.skipped_count} skipped, {result.aborted_count} aborted")
print(f"Aborted: {result.aborted}")
print()

for step_name, step_result in result.results.items():
    status_icon = {"success": "+", "failed": "X", "skipped": "-"}.get(step_result.status, "?")
    print(f"[{status_icon}] {step_name}: {step_result.status}")
    if step_result.status == "success":
        print(f"    Model: {step_result.model}")
        if step_result.usage:
            print(f"    Tokens: {step_result.usage.input_tokens} in + {step_result.usage.output_tokens} out")
        resp_str = str(step_result.response)
        print(f"    Response: {resp_str[:200]}{'...' if len(resp_str) > 200 else ''}")
    print()
""")


md("""\
---
<div class="page-break"></div>

## Section 4: Write Results

Save the execution results to an Excel file.
""")


code("""\
saved = write_workflow_results_excel(result, path=RESULTS_FILE, sheet="Results", spec=spec)
print(f"Saved to {saved}")

# Read back and display
wb_results = openpyxl.load_workbook(saved)
ws_results = wb_results["Results"]
for row in ws_results.iter_rows(values_only=True):
    print("  ".join(str(c) if c is not None else "" for c in row))
""")


md("""\
---
<div class="page-break"></div>

## How Conditions Work

### `condition` — Skip a step unless true

The `revise` step uses:
```
json_get({{score.response}}, "pass") == False
```
The `score` step returns JSON like `{"score": 7, "pass": true}`. The `json_get`
function extracts the `"pass"` field. If it's `False` (score < 8), `revise` runs.
If `pass` is `True`, `revise` is **skipped**.

### `abort_condition` — Stop the workflow if true

The `safety` step uses:
```
json_get({{safety.response}}, "flagged") == True
```
The `safety` step returns JSON like `{"safe": true, "flagged": false}`. If the
`flagged` field is `True`, the entire workflow **aborts**. No further steps run.
The `summary` step only executes if the content passes safety review.

### Condition syntax reference

| Expression | Meaning |
|-----------|---------|
| `{{step.response}} contains "text"` | Response includes substring |
| `{{step.response}} not contains "text"` | Response excludes substring |
| `{{step.status}} == "success"` | Status check |
| `{{step.has_response}}` | Has a non-empty response |
| `len({{step.response}}) > 100` | Response length check |
| `{{a.status}} == "success" and {{b.status}} == "success"` | Boolean AND |
| `{{step.response}} matches "pattern"` | Regex match |

### Next steps

- Try changing the `condition` threshold from 8 to a different value
- Add `abort_condition` to other steps for early termination
- Configure **passthrough columns** and **extra output columns** in `config/adapters.yaml`
- See [Excel adapter docs](../../docs/excel.md) for full configuration options
""")


out_path = Path(__file__).resolve().parent / "excel_workflow.ipynb"
with open(out_path, "w") as f:
    nbf.write(nb, f)

print(f"Created {out_path}")
