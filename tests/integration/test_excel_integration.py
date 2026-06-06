"""Excel integration tests — real workbook, real workflow execution, real write.

Run with: pytest tests/integration/test_excel_integration.py -m excel

Requires: MISTRAL_API_KEY (and OPENAI_API_KEY for GPT steps)
"""

from __future__ import annotations

import asyncio
from pathlib import Path

import pytest

from ffai_workflow_adapters.excel import load_workflow_excel, write_workflow_results_excel


def _create_client():
    import os

    from dotenv import load_dotenv
    from ffai import FFAI
    from ffai.Clients.AsyncFFLiteLLMClient import AsyncFFLiteLLMClient

    load_dotenv()

    from ffai_workflow_adapters.config import get_config

    config = get_config()
    default_name = config.clients.default_client
    client_cfg = config.clients.get_client_type(default_name)
    assert client_cfg is not None

    model_string = f"{client_cfg.provider_prefix}{client_cfg.default_model}"
    api_key = os.environ.get(client_cfg.api_key_env, "")

    return FFAI(AsyncFFLiteLLMClient(model_string=model_string, api_key=api_key))


def _create_workflow_workbook(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Workflow"
    ws.append(["name", "prompt", "client", "history", "temperature", "max_tokens", "Comments", "Priority"])
    ws.append(["topic", "Name a famous scientific discovery in one sentence.", "litellm-mistral-small", None, 0.7, 256, "Generate topic", "High"])
    ws.append(["explain", "Briefly explain the impact of: {{topic.response}}", "litellm-mistral-small", "topic", 0.5, 256, "Expand on it", "Medium"])
    wb.save(path)
    wb.close()


@pytest.mark.integration
@pytest.mark.excel
class TestExcelIntegration:
    def test_load_execute_write_roundtrip(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        workbook = tmp_path / "workflow.xlsx"
        results = tmp_path / "results.xlsx"
        _create_workflow_workbook(workbook)

        spec = load_workflow_excel(str(workbook), name="integration_test")
        assert spec.name == "integration_test"
        assert len(spec.prompts) == 2
        assert spec.prompts[0].name == "topic"
        assert spec.prompts[1].history == ["topic"]

        ffai = _create_client()
        result = asyncio.run(ffai.workflow.execute_workflow(spec))

        assert result.success_count == 2
        assert result.failed_count == 0
        assert "topic" in result.results
        assert "explain" in result.results
        assert result.results["topic"].response
        assert result.results["explain"].response

        path = write_workflow_results_excel(result, path=str(results), spec=spec)
        assert Path(path).exists()

        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["Results"]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3  # header + 2 data rows

        header = list(rows[0])
        assert "Comments" in header
        assert "Priority" in header
        assert "run_id" in header
        assert "run_date" in header

        comments_col = header.index("Comments")
        assert rows[1][comments_col] == "Generate topic"
        assert rows[2][comments_col] == "Expand on it"

        priority_col = header.index("Priority")
        assert rows[1][priority_col] == "High"
        assert rows[2][priority_col] == "Medium"

        run_id_col = header.index("run_id")
        assert rows[1][run_id_col] == rows[2][run_id_col]

        wb.close()

    def test_append_second_run(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        workbook = tmp_path / "workflow.xlsx"
        results = tmp_path / "results.xlsx"
        _create_workflow_workbook(workbook)

        spec = load_workflow_excel(str(workbook), name="append_test")
        ffai = _create_client()

        result1 = asyncio.run(ffai.workflow.execute_workflow(spec))
        write_workflow_results_excel(result1, path=str(results), spec=spec)

        result2 = asyncio.run(ffai.workflow.execute_workflow(spec))
        write_workflow_results_excel(result2, path=str(results), spec=spec)

        from openpyxl import load_workbook
        wb = load_workbook(results)
        ws = wb["Results"]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 5  # header + 2 + 2

        run_id_col = list(rows[0]).index("run_id")
        run_id_1 = rows[1][run_id_col]
        run_id_2 = rows[3][run_id_col]
        assert run_id_1 != run_id_2

        wb.close()

    def test_named_adapter_with_field_mapping(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        from openpyxl import Workbook

        reload_config()

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Steps"
        ws.append(["Task", "Instructions", "AI Model", "Context", "Comments"])
        ws.append(["topic", "Name a famous invention in one sentence.", "litellm-mistral-small", None, "Invention step"])
        wb.save(tmp_path / "custom.xlsx")
        wb.close()

        spec = load_workflow_excel(str(tmp_path / "custom.xlsx"), adapter="custom", sheet="Steps", name="custom_integration")
        assert spec.prompts[0].name == "topic"

        meta = getattr(spec, "_source_metadata", None)
        assert meta is not None
        assert meta["topic"]["Comments"] == "Invention step"

        ffai = _create_client()
        result = asyncio.run(ffai.workflow.execute_workflow(spec))
        assert result.success_count == 1

        results_path = tmp_path / "custom_results.xlsx"
        write_workflow_results_excel(result, path=str(results_path), adapter="custom", spec=spec)

        from openpyxl import load_workbook as lb
        wb2 = lb(results_path)
        ws2 = wb2["Custom Results"]
        rows = list(ws2.iter_rows(values_only=True))
        header = list(rows[0])
        assert "Task" in header
        assert "Output" in header
        assert "Comments" in header
        assert "run_id" in header

        task_col = header.index("Task")
        assert rows[1][task_col] == "topic"

        output_col = header.index("Output")
        assert rows[1][output_col]

        wb2.close()

    def test_custom_run_id(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        workbook = tmp_path / "workflow.xlsx"
        results = tmp_path / "results.xlsx"
        _create_workflow_workbook(workbook)

        spec = load_workflow_excel(str(workbook), name="custom_run_id")
        ffai = _create_client()
        result = asyncio.run(ffai.workflow.execute_workflow(spec))

        write_workflow_results_excel(result, path=str(results), spec=spec, run_id="batch-99")

        from openpyxl import load_workbook
        wb = load_workbook(results)
        ws = wb["Results"]
        rows = list(ws.iter_rows(values_only=True))
        run_id_col = list(rows[0]).index("run_id")
        assert rows[1][run_id_col] == "batch-99"
        assert rows[2][run_id_col] == "batch-99"
        wb.close()

    def test_config_output_path(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        workbook = tmp_path / "workflow.xlsx"
        _create_workflow_workbook(workbook)

        spec = load_workflow_excel(str(workbook), name="config_path_test")
        ffai = _create_client()
        result = asyncio.run(ffai.workflow.execute_workflow(spec))

        config_results = tmp_path / "auto_results.xlsx"
        from ffai_workflow_adapters.config import get_config
        cfg = get_config()
        saved_path = cfg.adapters.excel.output_path
        cfg.adapters.excel.output_path = str(config_results)
        try:
            write_workflow_results_excel(result, spec=spec)
            assert config_results.exists()
        finally:
            cfg.adapters.excel.output_path = saved_path
