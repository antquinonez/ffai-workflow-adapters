from __future__ import annotations

import logging
from pathlib import Path
from unittest.mock import patch

import pytest

from ffai.workflow.tabular import TabularLoadError
from ffai_workflow_adapters.excel import load_workflow_excel, write_workflow_results_excel, _resolve_extra_value


def _create_xlsx(path: Path, headers: list[str], rows: list[list], sheet: str | None = None) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    if sheet:
        ws.title = sheet
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


class TestResolveExtraValue:
    def test_now_template_custom_format(self):
        import re
        result = _resolve_extra_value("{{now:%Y/%m/%d}}", "run-1")
        assert re.match(r"\d{4}/\d{2}/\d{2}", result)

    def test_now_template_time_format(self):
        import re
        result = _resolve_extra_value("{{now:%H:%M:%S}}", "run-1")
        assert re.match(r"\d{2}:\d{2}:\d{2}", result)

    def test_literal_string_passthrough(self):
        result = _resolve_extra_value("static-value", "run-1")
        assert result == "static-value"

    def test_run_id_template(self):
        result = _resolve_extra_value("{{run_id}}", "batch-42")
        assert result == "batch-42"

    def test_date_template(self):
        import re
        result = _resolve_extra_value("{{date}}", "run-1")
        assert re.match(r"\d{4}-\d{2}-\d{2}", result)

    def test_timestamp_template(self):
        result = _resolve_extra_value("{{timestamp}}", "run-1")
        assert "T" in result


class TestLoadWorkflowExcel:
    def test_basic_load(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        filepath = tmp_path / "workflow.xlsx"
        _create_xlsx(filepath, ["name", "prompt", "history"], [
            ["topic", "Name a discovery."],
            ["explain", "Explain {{topic.response}}.", "topic"],
        ])

        spec = load_workflow_excel(filepath, name="test")
        assert spec.name == "test"
        assert len(spec.prompts) == 2
        assert spec.prompts[0].name == "topic"
        assert spec.prompts[1].history == ["topic"]

    def test_specific_sheet_by_name(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        filepath = tmp_path / "workflow.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Other"
        ws1.append(["x", "y"])
        ws1.append(["a", "b"])
        ws2 = wb.create_sheet("Steps")
        ws2.append(["name", "prompt"])
        ws2.append(["topic", "Go"])
        wb.save(filepath)
        wb.close()

        spec = load_workflow_excel(filepath, sheet="Steps", name="sheet_test")
        assert spec.name == "sheet_test"
        assert len(spec.prompts) == 1
        assert spec.prompts[0].name == "topic"

    def test_specific_sheet_by_index(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        filepath = tmp_path / "workflow.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "Skip"
        ws1.append(["x"])
        ws1.append(["skip"])
        ws2 = wb.create_sheet("Steps")
        ws2.append(["name", "prompt"])
        ws2.append(["topic", "Go"])
        wb.save(filepath)
        wb.close()

        spec = load_workflow_excel(filepath, sheet=1, name="index_test")
        assert len(spec.prompts) == 1

    def test_with_clients(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        filepath = tmp_path / "workflow.xlsx"
        _create_xlsx(filepath, ["name", "prompt", "client"], [
            ["step1", "Go", "reviewer"],
        ])

        spec = load_workflow_excel(
            filepath,
            clients={"reviewer": {"type": "litellm", "model": "gpt-4o"}},
        )
        assert spec.prompts[0].client is not None
        assert spec.prompts[0].client.name == "reviewer"

    def test_with_defaults(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        filepath = tmp_path / "workflow.xlsx"
        _create_xlsx(filepath, ["name", "prompt"], [
            ["step1", "Go"],
        ])

        spec = load_workflow_excel(
            filepath,
            defaults={"temperature": 0.5, "max_tokens": 200},
        )
        assert spec.defaults.temperature == 0.5
        assert spec.defaults.max_tokens == 200

    def test_empty_file_raises(self, tmp_path: Path):
        filepath = tmp_path / "empty.xlsx"
        _create_xlsx(filepath, ["name", "prompt"], [])

        with pytest.raises(TabularLoadError, match="no data rows"):
            load_workflow_excel(filepath)

    def test_missing_file_raises(self, tmp_path: Path):
        with pytest.raises(TabularLoadError, match="not found"):
            load_workflow_excel(tmp_path / "nonexistent.xlsx")

    def test_missing_openpyxl_raises(self, tmp_path: Path):
        filepath = tmp_path / "workflow.xlsx"
        _create_xlsx(filepath, ["name", "prompt"], [["a", "Go"]])

        with patch.dict("sys.modules", {"openpyxl": None}):
            with pytest.raises(TabularLoadError, match="openpyxl is required"):
                load_workflow_excel(filepath)

    def test_input_field_mapping(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        cfg.adapters.excel.input_field_map = {"Task": "name", "Instructions": "prompt"}
        try:
            filepath = tmp_path / "workflow.xlsx"
            _create_xlsx(filepath, ["Task", "Instructions"], [
                ["topic", "Name a discovery."],
            ])

            spec = load_workflow_excel(filepath, name="mapped")
            assert spec.prompts[0].name == "topic"
            assert spec.prompts[0].prompt == "Name a discovery."
        finally:
            cfg.adapters.excel.input_field_map = {}

    def test_passthrough_columns(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        cfg.adapters.excel.input_field_map = {}
        cfg.adapters.excel.passthrough_columns = ["Comments", "Priority"]
        try:
            filepath = tmp_path / "workflow.xlsx"
            _create_xlsx(filepath, ["name", "prompt", "Comments", "Priority"], [
                ["topic", "Go", "Check refs", "High"],
                ["explain", "Expand", None, "Low"],
            ])

            spec = load_workflow_excel(filepath, name="passthrough_test")
            meta = getattr(spec, "_source_metadata", None)
            assert meta is not None
            assert meta["topic"]["Comments"] == "Check refs"
            assert meta["topic"]["Priority"] == "High"
            assert meta["explain"]["Comments"] is None
            assert meta["explain"]["Priority"] == "Low"
        finally:
            cfg.adapters.excel.input_field_map = {}
            cfg.adapters.excel.passthrough_columns = []

    def test_passthrough_with_field_mapping(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        cfg.adapters.excel.input_field_map = {"Task": "name", "Instructions": "prompt"}
        cfg.adapters.excel.passthrough_columns = ["Notes"]
        try:
            filepath = tmp_path / "workflow.xlsx"
            _create_xlsx(filepath, ["Task", "Instructions", "Notes"], [
                ["step1", "Go", "Important note"],
            ])

            spec = load_workflow_excel(filepath, name="mapped_passthrough")
            assert spec.prompts[0].name == "step1"
            meta = getattr(spec, "_source_metadata", None)
            assert meta is not None
            assert meta["step1"]["Notes"] == "Important note"
        finally:
            cfg.adapters.excel.input_field_map = {}
            cfg.adapters.excel.passthrough_columns = []

    def test_no_passthrough_without_config(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        cfg.adapters.excel.input_field_map = {}
        cfg.adapters.excel.passthrough_columns = []
        try:
            filepath = tmp_path / "workflow.xlsx"
            _create_xlsx(filepath, ["name", "prompt", "Comments"], [
                ["step1", "Go", "Some comment"],
            ])

            spec = load_workflow_excel(filepath, name="no_pt")
            assert not hasattr(spec, "_source_metadata")
        finally:
            pass

    def test_schema_missing_required_column_raises(self, tmp_path: Path):
        filepath = tmp_path / "workflow.xlsx"
        _create_xlsx(filepath, ["name_only", "other"], [
            ["step1", "value"],
        ])

        with pytest.raises(TabularLoadError, match="missing required columns"):
            load_workflow_excel(filepath)

    def test_schema_missing_prompt_with_field_map_raises(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        cfg.adapters.excel.input_field_map = {"Task": "name"}
        try:
            filepath = tmp_path / "workflow.xlsx"
            _create_xlsx(filepath, ["Task", "Other"], [
                ["step1", "value"],
            ])

            with pytest.raises(TabularLoadError, match="missing required columns"):
                load_workflow_excel(filepath)
        finally:
            cfg.adapters.excel.input_field_map = {}

    def test_schema_invalid_temperature_raises(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        filepath = tmp_path / "workflow.xlsx"
        _create_xlsx(filepath, ["name", "prompt", "temperature"], [
            ["step1", "Go", "hot"],
        ])

        with pytest.raises(TabularLoadError, match="temperature must be a number"):
            load_workflow_excel(filepath)

    def test_schema_invalid_max_tokens_raises(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        filepath = tmp_path / "workflow.xlsx"
        _create_xlsx(filepath, ["name", "prompt", "max_tokens"], [
            ["step1", "Go", "lots"],
        ])

        with pytest.raises(TabularLoadError, match="max_tokens must be a number"):
            load_workflow_excel(filepath)

    def test_schema_valid_numeric_fields_pass(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        reload_config()

        filepath = tmp_path / "workflow.xlsx"
        _create_xlsx(filepath, ["name", "prompt", "temperature", "max_tokens"], [
            ["step1", "Go", 0.7, 1024],
        ])

        spec = load_workflow_excel(filepath, name="valid_nums")
        assert spec.prompts[0].name == "step1"

    def test_schema_warns_on_unrecognized_columns(self, tmp_path: Path, caplog):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        cfg.adapters.excel.input_field_map = {}
        cfg.adapters.excel.passthrough_columns = []
        try:
            filepath = tmp_path / "workflow.xlsx"
            _create_xlsx(filepath, ["name", "prompt", "WackyColumn", "AnotherBad"], [
                ["step1", "Go", "x", "y"],
            ])

            with caplog.at_level(logging.WARNING):
                spec = load_workflow_excel(filepath, name="warn_test")
            assert spec.prompts[0].name == "step1"
            assert "Unrecognized columns" in caplog.text
            assert "WackyColumn" in caplog.text
        finally:
            pass


class TestWriteWorkflowResultsExcel:
    def _make_result(self):
        from ffai.core.response_result import ResponseResult
        from ffai.core.usage import TokenUsage
        from dataclasses import dataclass, field

        @dataclass
        class FakeWorkflowResult:
            results: dict = field(default_factory=dict)
            success_count: int = 0
            failed_count: int = 0
            skipped_count: int = 0
            aborted: bool = False
            aborted_count: int = 0
            spec_name: str = "test_workflow"

        return FakeWorkflowResult(
            success_count=2,
            results={
                "topic": ResponseResult(
                    response="DNA discovered.",
                    model="mistral-small-latest",
                    status="success",
                    usage=TokenUsage(input_tokens=41, output_tokens=40, total_tokens=81),
                    duration_ms=1234.5,
                ),
                "explain": ResponseResult(
                    response="It changed everything.",
                    model="gpt-4o-mini",
                    status="success",
                    usage=TokenUsage(input_tokens=84, output_tokens=161, total_tokens=245),
                    cost_usd=0.002,
                    duration_ms=2345.6,
                ),
            },
        )

    def test_write_new_file(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved = dict(cfg.adapters.excel.output_field_map)
        cfg.adapters.excel.output_field_map = {}
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()

            path = write_workflow_results_excel(result, path=filepath)
            assert Path(path).exists()

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            assert rows[0][0] == "workflow"
            assert rows[0][1] == "step"
            assert len(rows) == 3  # header + 2 results
            assert rows[1][1] == "topic"
            assert rows[2][1] == "explain"
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = saved

    def test_write_to_existing_file(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved = dict(cfg.adapters.excel.output_field_map)
        cfg.adapters.excel.output_field_map = {}
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()

            write_workflow_results_excel(result, path=filepath)
            write_workflow_results_excel(result, path=filepath, sheet="Run2")

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            assert "Results" in wb.sheetnames
            assert "Run2" in wb.sheetnames
            ws1 = wb["Results"]
            ws2 = wb["Run2"]
            assert list(ws1.iter_rows(values_only=True))[0][1] == "step"
            assert list(ws2.iter_rows(values_only=True))[0][1] == "step"
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = saved

    def test_append_to_existing_sheet(self, tmp_path: Path):
        filepath = tmp_path / "results.xlsx"
        result = self._make_result()

        write_workflow_results_excel(result, path=filepath)
        write_workflow_results_excel(result, path=filepath)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb["Results"]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 5  # header + 2 + 2
        wb.close()

    def test_output_field_mapping(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        cfg.adapters.excel.output_field_map = {"step": "Task", "response": "Output"}
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()

            write_workflow_results_excel(result, path=filepath)

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            header = list(rows[0])
            assert "Task" in header
            assert "Output" in header
            assert "step" not in header
            assert "response" not in header
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = {}

    def test_config_output_path(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved_path = cfg.adapters.excel.output_path
        saved_sheet = cfg.adapters.excel.output_sheet
        saved_map = dict(cfg.adapters.excel.output_field_map)
        cfg.adapters.excel.output_field_map = {}
        cfg.adapters.excel.output_path = str(tmp_path / "config_results.xlsx")
        cfg.adapters.excel.output_sheet = "AutoResults"
        try:
            result = self._make_result()
            path = write_workflow_results_excel(result)
            assert Path(path).exists()

            from openpyxl import load_workbook
            wb = load_workbook(path)
            assert "AutoResults" in wb.sheetnames
            wb.close()
        finally:
            cfg.adapters.excel.output_path = saved_path
            cfg.adapters.excel.output_sheet = saved_sheet
            cfg.adapters.excel.output_field_map = saved_map

    def test_no_path_raises(self):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved_path = cfg.adapters.excel.output_path
        cfg.adapters.excel.output_path = ""
        try:
            result = self._make_result()
            with pytest.raises(ValueError, match="No output path"):
                write_workflow_results_excel(result)
        finally:
            cfg.adapters.excel.output_path = saved_path

    def test_override_sheet_from_config(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved_map = dict(cfg.adapters.excel.output_field_map)
        cfg.adapters.excel.output_field_map = {}
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()
            write_workflow_results_excel(result, path=filepath, sheet="Custom")

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            assert "Custom" in wb.sheetnames
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = saved_map

    def test_run_id_auto_generated(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        import re
        cfg = reload_config()
        saved_map = dict(cfg.adapters.excel.output_field_map)
        saved_extra = dict(cfg.adapters.excel.extra_output_columns)
        cfg.adapters.excel.output_field_map = {}
        cfg.adapters.excel.extra_output_columns = {"run_id": "{{run_id}}"}
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()

            write_workflow_results_excel(result, path=filepath)

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            header = list(rows[0])
            run_id_col = header.index("run_id")
            run_id_1 = rows[1][run_id_col]
            run_id_2 = rows[2][run_id_col]
            assert run_id_1 == run_id_2
            assert re.match(r"\d{8}-\d{6}", str(run_id_1))
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = saved_map
            cfg.adapters.excel.extra_output_columns = saved_extra

    def test_run_id_custom(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved_map = dict(cfg.adapters.excel.output_field_map)
        saved_extra = dict(cfg.adapters.excel.extra_output_columns)
        cfg.adapters.excel.output_field_map = {}
        cfg.adapters.excel.extra_output_columns = {"run_id": "{{run_id}}"}
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()

            write_workflow_results_excel(result, path=filepath, run_id="batch-42")

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            header = list(rows[0])
            run_id_col = header.index("run_id")
            assert rows[1][run_id_col] == "batch-42"
            assert rows[2][run_id_col] == "batch-42"
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = saved_map
            cfg.adapters.excel.extra_output_columns = saved_extra

    def test_run_id_consistent_across_rows(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved_map = dict(cfg.adapters.excel.output_field_map)
        saved_extra = dict(cfg.adapters.excel.extra_output_columns)
        cfg.adapters.excel.output_field_map = {}
        cfg.adapters.excel.extra_output_columns = {
            "run_id": "{{run_id}}",
            "ts": "{{timestamp}}",
        }
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()

            write_workflow_results_excel(result, path=filepath)

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            header = list(rows[0])
            run_id_col = header.index("run_id")
            ts_col = header.index("ts")
            assert rows[1][run_id_col] == rows[2][run_id_col]
            assert rows[1][ts_col] == rows[2][ts_col]
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = saved_map
            cfg.adapters.excel.extra_output_columns = saved_extra

    def test_extra_output_columns(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved_map = dict(cfg.adapters.excel.output_field_map)
        saved_extra = dict(cfg.adapters.excel.extra_output_columns)
        cfg.adapters.excel.output_field_map = {}
        cfg.adapters.excel.extra_output_columns = {"run_date": "{{date}}", "batch": "test-run-01"}
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()

            write_workflow_results_excel(result, path=filepath)

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            header = list(rows[0])
            assert "run_date" in header
            assert "batch" in header
            batch_col = header.index("batch")
            assert rows[1][batch_col] == "test-run-01"
            assert rows[2][batch_col] == "test-run-01"
            run_date = rows[1][header.index("run_date")]
            assert isinstance(run_date, str)
            assert len(run_date) == 10
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = saved_map
            cfg.adapters.excel.extra_output_columns = saved_extra

    def test_passthrough_and_extra_together(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved_map = dict(cfg.adapters.excel.output_field_map)
        saved_pt = list(cfg.adapters.excel.passthrough_columns)
        saved_extra = dict(cfg.adapters.excel.extra_output_columns)
        cfg.adapters.excel.output_field_map = {}
        cfg.adapters.excel.passthrough_columns = ["Notes"]
        cfg.adapters.excel.extra_output_columns = {"run_date": "{{date}}"}
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()

            from ffai.workflow.tabular import load_workflow_rows
            spec = load_workflow_rows(
                [{"name": "topic", "prompt": "Go"}],
                name="test",
            )
            object.__setattr__(spec, "_source_metadata", {"topic": {"Notes": "see above"}})

            write_workflow_results_excel(result, path=filepath, spec=spec)

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            header = list(rows[0])
            assert "Notes" in header
            assert "run_date" in header
            assert header.index("Notes") < header.index("run_date")
            notes_col = header.index("Notes")
            assert rows[1][notes_col] == "see above"
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = saved_map
            cfg.adapters.excel.passthrough_columns = saved_pt
            cfg.adapters.excel.extra_output_columns = saved_extra

    def test_write_without_spec_no_passthrough(self, tmp_path: Path):
        from ffai_workflow_adapters.config import reload_config
        cfg = reload_config()
        saved_map = dict(cfg.adapters.excel.output_field_map)
        saved_pt = list(cfg.adapters.excel.passthrough_columns)
        cfg.adapters.excel.output_field_map = {}
        cfg.adapters.excel.passthrough_columns = ["Comments"]
        try:
            filepath = tmp_path / "results.xlsx"
            result = self._make_result()

            write_workflow_results_excel(result, path=filepath)

            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb["Results"]
            rows = list(ws.iter_rows(values_only=True))
            header = list(rows[0])
            assert "Comments" in header
            comments_col = header.index("Comments")
            assert rows[1][comments_col] is None
            wb.close()
        finally:
            cfg.adapters.excel.output_field_map = saved_map
            cfg.adapters.excel.passthrough_columns = saved_pt

    def test_write_missing_openpyxl_raises(self, tmp_path: Path):
        result = self._make_result()
        filepath = tmp_path / "results.xlsx"

        with patch.dict("sys.modules", {"openpyxl": None}):
            with pytest.raises(TabularLoadError, match="openpyxl is required"):
                write_workflow_results_excel(result, path=filepath)
