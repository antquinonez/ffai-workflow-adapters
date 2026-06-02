from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from ffai.workflow.tabular import TabularLoadError, load_workflow_rows

from ffai_workflow_adapters._templates import _generate_run_id, _resolve_extra_value
from ffai_workflow_adapters._validation import validate_schema
from ffai_workflow_adapters.config import get_config

logger = logging.getLogger(__name__)

CANONICAL_FIELDS = frozenset({
    "name", "prompt", "client", "model", "history",
    "temperature", "max_tokens", "system_instructions",
    "condition", "abort_condition", "response_format",
    "response_model", "strict", "tools", "tool_choice",
})


def load_workflow_excel(
    path: str | Path,
    *,
    sheet: str | int | None = None,
    adapter: str | None = None,
    name: str = "unnamed",
    description: str = "",
    defaults: dict[str, Any] | None = None,
    clients: dict[str, dict[str, Any] | str] | None = None,
    tools: dict[str, dict[str, Any]] | None = None,
) -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise TabularLoadError(
            "openpyxl is required for Excel loading. Install with: pip install ffai-workflow-adapters[excel]"
        ) from e

    filepath = Path(path)
    if not filepath.exists():
        raise TabularLoadError(f"Excel file not found: {filepath}")

    wb = load_workbook(filepath, read_only=True, data_only=True)

    if sheet is None:
        ws = wb.active
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]
    assert ws is not None

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        wb.close()
        raise TabularLoadError(f"Excel file '{filepath}' has no header row")

    headers = [str(h).strip() if h is not None else "" for h in header]

    cfg = get_config().adapters.excel.resolve(adapter)
    field_map = cfg.input_field_map or None
    passthrough = cfg.passthrough_columns or []

    valid_columns = CANONICAL_FIELDS | set(passthrough)
    if field_map:
        recognized = set(field_map.keys()) | valid_columns
        unmapped = [h for h in headers if h and h not in recognized]
    else:
        unmapped = [h for h in headers if h and h not in valid_columns]
    if unmapped:
        logger.warning(
            "Unrecognized columns in '%s' (sheet '%s'): %s. "
            "These will be passed through unless they match a canonical field.",
            filepath,
            getattr(ws, "title", "?"),
            unmapped,
        )

    rows: list[dict[str, Any]] = []
    source_metadata: dict[str, dict[str, Any]] = {}

    for row in rows_iter:
        record: dict[str, Any] = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = val
        if any(v is not None for v in record.values()):
            if field_map:
                mapped: dict[str, Any] = {}
                for col, val in record.items():
                    canonical = field_map.get(col, col)
                    if canonical is not None:
                        mapped[canonical] = val
                rows.append(mapped)
                step_name = str(mapped.get("name", ""))
            else:
                rows.append(record)
                step_name = str(record.get("name", ""))

            if passthrough and step_name:
                pt_data = {col: record[col] for col in passthrough if col in record}
                if pt_data:
                    source_metadata[step_name] = pt_data

    wb.close()

    if not rows:
        raise TabularLoadError(f"Excel file '{filepath}' contains no data rows")

    validate_schema(rows, str(filepath))

    spec = load_workflow_rows(
        rows,
        name=name,
        description=description,
        defaults=defaults,
        clients=clients,
        tools=tools,
    )

    if source_metadata:
        object.__setattr__(spec, "_source_metadata", source_metadata)

    return spec


def write_workflow_results_excel(
    result: Any,
    path: str | Path | None = None,
    *,
    sheet: str | None = None,
    adapter: str | None = None,
    spec: Any | None = None,
    run_id: str | None = None,
) -> str:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as e:
        raise TabularLoadError(
            "openpyxl is required for Excel output. Install with: pip install ffai-workflow-adapters[excel]"
        ) from e

    cfg = get_config().adapters.excel.resolve(adapter)
    output_map = cfg.output_field_map

    filepath = Path(path) if path else Path(cfg.output_path) if cfg.output_path else None
    if not filepath:
        raise ValueError("No output path specified: pass path= or set output_path in config")

    sheet_name = sheet or cfg.output_sheet or "Results"

    source_metadata = getattr(spec, "_source_metadata", None) if spec else None

    passthrough_cols = cfg.passthrough_columns or []
    extra_cols = cfg.extra_output_columns or {}

    resolved_run_id = run_id or _generate_run_id()
    resolved_extras = {col: _resolve_extra_value(tmpl, resolved_run_id) for col, tmpl in extra_cols.items()}

    canonical_fields = [
        "workflow", "step", "status", "response", "model",
        "input_tokens", "output_tokens", "cost_usd", "duration_ms", "timestamp",
    ]

    timestamp = _resolve_extra_value("{{timestamp}}", resolved_run_id)
    records: list[dict[str, Any]] = []

    for step_name, step_result in result.results.items():
        fields: dict[str, Any] = {
            "workflow": result.spec_name or "",
            "step": step_name,
            "status": step_result.status,
            "response": step_result.response or "",
            "model": step_result.model or "",
            "timestamp": timestamp,
        }
        if step_result.usage:
            fields["input_tokens"] = step_result.usage.input_tokens
            fields["output_tokens"] = step_result.usage.output_tokens
        if step_result.cost_usd:
            fields["cost_usd"] = step_result.cost_usd
        if step_result.duration_ms:
            fields["duration_ms"] = round(step_result.duration_ms, 1)

        if source_metadata and step_name in source_metadata:
            for col in passthrough_cols:
                if col in source_metadata[step_name]:
                    fields[col] = source_metadata[step_name][col]

        for col_name, value in resolved_extras.items():
            fields[col_name] = value

        if output_map:
            fields = {output_map.get(k, k): v for k, v in fields.items()}

        records.append(fields)

    passthrough_headers = [output_map.get(c, c) if output_map else c for c in passthrough_cols]
    extra_headers = [output_map.get(c, c) if output_map else c for c in extra_cols]
    column_names = (
        [output_map.get(f, f) if output_map else f for f in canonical_fields]
        + passthrough_headers
        + extra_headers
    )

    if filepath.exists():
        wb = load_workbook(filepath)
    else:
        filepath.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws_active = wb.active
        assert ws_active is not None
        ws_active.title = sheet_name

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    has_header = any(ws.cell(1, col).value is not None for col in range(1, len(column_names) + 1))
    if not has_header:
        for col_idx, col_name in enumerate(column_names, start=1):
            ws.cell(1, col_idx, col_name)

    for rec in records:
        ws.append([rec.get(col) for col in column_names])

    wb.save(filepath)
    wb.close()
    return str(filepath)
